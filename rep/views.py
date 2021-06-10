from django.contrib.auth.mixins import PermissionRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from matr.models import *


class ReporteCurso(PermissionRequiredMixin, TemplateView):
    # Clase para generar el reporte de un grupo de estudiantes
    template_name = "rep_curso.html"
    permission_required = "rep.view_curso"

    def get_context_data(self, *, object_list=None, **kwargs):
        # Con este metodo realizo los filtros de los campos
        context = super().get_context_data(**kwargs)
        context["al"] = AnioLectivo.objects.all()
        context["jo"] = GenrGeneral.objects.filter(tipo="JOR")
        return context


class ReporteMatricula(PermissionRequiredMixin, TemplateView):
    # Clase para generar el reporte de matricula de un estudiante
    template_name = "rep_matricula.html"
    permission_required = "rep.view_matricula"

    def get_context_data(self, *, object_list=None, **kwargs):
        # Con este metodo realizo los filtros de los campos
        context = super().get_context_data(**kwargs)
        context["al"] = AnioLectivo.objects.all()
        context["jo"] = GenrGeneral.objects.filter(tipo="JOR")
        context["es"] = Persona.objects.filter(is_estudiante=True)
        return context


def con_jor(request):
    # Consultar jornada
    var = list()
    if request.method == "POST":
        jor_id = request.POST.get("jornada")
        j = GenrGeneral.objects.filter(idgenr_general=jor_id).first()
        if j:
            curso = CabCurso.objects.filter(id_cfg_jornada=j)
            for i in curso:
                var.append({"id": i.id_curso, "curso": i.nombre})
        return JsonResponse({"lista": var}, status=200, )


def con_cur(request):
    # Consultar curso
    var_v = list()
    if request.method == "POST":
        cur_id = request.POST.get("curso_i")
        k = CabCurso.objects.filter(id_curso=cur_id).first()
        if k:
            seccion = Aniolectivo_curso.objects.filter(id_curso=k)
            for l in seccion:
                var_v.append({"id_c": l.id_matr_anioelectivo_curso, "paralelo": l.paralelo})
        return JsonResponse({"listaa": var_v}, status=200, )


class ReporteProfesor(PermissionRequiredMixin, TemplateView):
    # Clase para generar el reporte de un Profesor
    template_name = "rep_profesores.html"
    permission_required = "rep.view_estudiante"

    def get_context_data(self, *, object_list=None, **kwargs):
        # Con este metodo realizo los filtros de los campos
        context = super().get_context_data(**kwargs)
        context["obj"] = Persona.objects.filter(is_empleado=True)
        return context

    def post(self, request):
        """ metodo para ejecutar el metodo post y descargue
        el archivo de excel"""
        id_prof = request.POST.get('id_per')
        if len(id_prof) == 0:
            return redirect("rep:reporte_profesores")
        profesor = Persona.objects.filter(id_persona=id_prof)
        return self.excel_profesor(profesor)

    def excel_profesor(self, n_prof):
        """Metodo para diseño de archivo de excel
        para el reporte de profesor"""
        wb = Workbook()
        ws = wb.active
        # Sección para crear hojas en excel
        ws.title = 'Información Personal'
        ws1 = wb.create_sheet("Historial Académico")
        # Sección Informacion Personal
        # Sección para añadir contenido a las celdas
        for prof in n_prof:
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A1'].font = Font(name='times new roman', size=14, bold=True)
            ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['A1'] = 'Información Personal'
            ws['A2'].border = Border(left=Side(border_style="thin"))
            ws['I2'].border = Border(right=Side(border_style="thin"))
            ws['A22'].border = Border(bottom=Side(border_style="thin"), right=Side(border_style="thin"),
                                      left=Side(border_style="thin"))
            ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B3'].font = Font(name='times new roman', size=12, bold=True)
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'] = 'Foto'
            ws['C3'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C3'].font = Font(name='times new roman', size=12, bold=True)
            ws['C3'] = 'Cédula:'
            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E3'].font = Font(name='times new roman', size=12)
            ws['E3'].border = Border(bottom=Side(border_style="thin"))
            ws['E3'] = prof.identificacion
            ws['C4'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C4'].font = Font(name='times new roman', size=12, bold=True)
            ws['C4'] = 'Nombres:'
            ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E4'].font = Font(name='times new roman', size=12)
            ws['E4'].border = Border(bottom=Side(border_style="thin"))
            n = prof.nombres + " " + prof.apellidos
            nombres = n.title()
            ws['E4'] = nombres
            ws['C5'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C5'].font = Font(name='times new roman', size=12, bold=True)
            ws['C5'] = 'Telefono:'
            ws['E5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E5'].font = Font(name='times new roman', size=12)
            ws['E5'].border = Border(bottom=Side(border_style="thin"))
            ws['E5'] = prof.telefono
            ws['C6'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C6'].font = Font(name='times new roman', size=12, bold=True)
            ws['C6'] = 'Celular:'
            ws['E6'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E6'].font = Font(name='times new roman', size=12)
            ws['E6'].border = Border(bottom=Side(border_style="thin"))
            ws['E6'] = prof.celular
            ws['C7'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C7'].font = Font(name='times new roman', size=12, bold=True)
            ws['C7'] = 'Correo:'
            ws['E7'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].font = Font(name='times new roman', size=12)
            ws['E7'].border = Border(bottom=Side(border_style="thin"))
            ws['E7'] = "Insertar correo"
            ws['C8'].alignment = Alignment(horizontal="right", vertical="center")
            ws['C8'].font = Font(name='times new roman', size=12, bold=True)
            ws['C8'] = 'Idioma Ancestral:'
            ws['E8'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E8'].font = Font(name='times new roman', size=12)
            ws['E8'].border = Border(bottom=Side(border_style="thin"))
            id_anc = prof.id_cfg_idioma_ancestral.nombre
            idioma = id_anc.title()
            ws['E8'] = idioma
            ws['B10'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B10'].font = Font(name='times new roman', size=12, bold=True)
            ws['B10'] = 'Fecha de nacimiento:'
            ws['D10'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D10'].font = Font(name='times new roman', size=11)
            ws['D10'].border = Border(bottom=Side(border_style="thin"))
            ws['D10'] = prof.fecha_de_nacimiento
            ws['B11'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B11'].font = Font(name='times new roman', size=12, bold=True)
            ws['B11'] = 'Lugar de nacimiento:'
            ws['D11'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D11'].font = Font(name='times new roman', size=11)
            ws['D11'].border = Border(bottom=Side(border_style="thin"))
            ws['D11'] = prof.lugar_nacimiento
            ws['B12'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B12'].font = Font(name='times new roman', size=12, bold=True)
            ws['B12'] = 'Estado civil:'
            ws['D12'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D12'].font = Font(name='times new roman', size=11)
            ws['D12'].border = Border(bottom=Side(border_style="thin"))
            ec = prof.id_cfg_estado_civil.nombre
            estado_c = ec.title()
            ws['D12'] = estado_c
            ws['B13'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B13'].font = Font(name='times new roman', size=12, bold=True)
            ws['B13'] = 'Categoria migratoria:'
            ws['D13'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D13'].font = Font(name='times new roman', size=11)
            ws['D13'].border = Border(bottom=Side(border_style="thin"))
            cm = prof.id_cfg_categoria_migratoria.nombre
            cat_mig = cm.title()
            ws['D13'] = cat_mig
            ws['E10'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E10'].font = Font(name='times new roman', size=12, bold=True)
            ws['E10'] = 'Dirección:'
            ws['F10'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F10'].font = Font(name='times new roman', size=11)
            ws['F10'].border = Border(bottom=Side(border_style="thin"))
            ws['F10'] = prof.direccion
            ws['E11'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E11'].font = Font(name='times new roman', size=12, bold=True)
            ws['E11'] = 'Pais:'
            ws['F11'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F11'].font = Font(name='times new roman', size=11)
            ws['F11'].border = Border(bottom=Side(border_style="thin"))
            p = prof.id_cfg_pais.nombre
            pais = p.title()
            ws['F11'] = pais
            ws['G11'].alignment = Alignment(horizontal="right", vertical="center")
            ws['G11'].font = Font(name='times new roman', size=12, bold=True)
            ws['G11'] = 'Provincia:'
            ws['H11'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H11'].font = Font(name='times new roman', size=11)
            ws['H11'].border = Border(bottom=Side(border_style="thin"))
            pro = prof.id_cfg_provincia.nombre
            provincia = pro.title()
            ws['H11'] = provincia
            ws['E12'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E12'].font = Font(name='times new roman', size=12, bold=True)
            ws['E12'] = 'Ciudad:'
            ws['F12'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F12'].font = Font(name='times new roman', size=11)
            ws['F12'].border = Border(bottom=Side(border_style="thin"))
            c = prof.id_cfg_ciudad.nombre
            ciudad = c.title()
            ws['F12'] = ciudad
            ws['G12'].alignment = Alignment(horizontal="right", vertical="center")
            ws['G12'].font = Font(name='times new roman', size=12, bold=True)
            ws['G12'] = 'Sexo:'
            ws['H12'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H12'].font = Font(name='times new roman', size=11)
            ws['H12'].border = Border(bottom=Side(border_style="thin"))
            s = prof.id_cfg_genero.nombre
            sexo = s.title()
            ws['H12'] = sexo
            ws['F13'].alignment = Alignment(horizontal="right", vertical="center")
            ws['F13'].font = Font(name='times new roman', size=12, bold=True)
            ws['F13'] = 'Etnia:'
            ws['G13'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G13'].font = Font(name='times new roman', size=11)
            ws['G13'].border = Border(bottom=Side(border_style="thin"))
            et = prof.id_cfg_etnia.nombre
            etnia = et.title()
            ws['G13'] = etnia
            ws['B15'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B15'].font = Font(name='times new roman', size=14, bold=True)
            ws['B15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B15'] = 'Documentación Médica'
            ws['E17'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E17'].font = Font(name='times new roman', size=12, bold=True)
            ws['E17'] = 'Discapacidad:'
            ws['G17'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G17'].font = Font(name='times new roman', size=11)
            ws['G17'].border = Border(bottom=Side(border_style="thin"))
            if prof.discapacidad:
                ws['G17'] = "Si"
            else:
                ws['G17'] = "No"
            ws['E18'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E18'].font = Font(name='times new roman', size=12, bold=True)
            ws['E18'] = 'Asma:'
            ws['G18'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G18'].font = Font(name='times new roman', size=11)
            ws['G18'].border = Border(bottom=Side(border_style="thin"))
            if prof.asma:
                ws['G18'] = "Si"
            else:
                ws['G18'] = "No"
            ws['E19'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E19'].font = Font(name='times new roman', size=12, bold=True)
            ws['E19'] = 'Epilepsia:'
            ws['G19'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G19'].font = Font(name='times new roman', size=11)
            ws['G19'].border = Border(bottom=Side(border_style="thin"))
            if prof.epilepsia:
                ws['G19'] = "Si"
            else:
                ws['G19'] = "No"
            ws['E20'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E20'].font = Font(name='times new roman', size=12, bold=True)
            ws['E20'] = 'Tipo de sangre:'
            ws['G20'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G20'].font = Font(name='times new roman', size=11)
            ws['G20'].border = Border(bottom=Side(border_style="thin"))
            ws['G20'] = prof.id_cfg_tipo_sangre.nombre
            ws['E21'].alignment = Alignment(horizontal="right", vertical="center")
            ws['E21'].font = Font(name='times new roman', size=12, bold=True)
            ws['E21'] = 'Alergias:'
            ws['G21'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G21'].font = Font(name='times new roman', size=11)
            ws['G21'].border = Border(bottom=Side(border_style="thin"))
            if prof.enfermedad_alergica:
                ws['G21'] = "Si"
            else:
                ws['G21'] = "No"
            ws['B17'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B17'].font = Font(name='times new roman', size=12, bold=True)
            ws['B17'] = 'Discapacidad renal:'
            ws['D17'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D17'].font = Font(name='times new roman', size=11)
            ws['D17'].border = Border(bottom=Side(border_style="thin"))
            if prof.discapacidad_renal:
                ws['D17'] = "Si"
            else:
                ws['D17'] = "No"
            ws['B18'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B18'].font = Font(name='times new roman', size=12, bold=True)
            ws['B18'] = 'Discapacidad neurologica:'
            ws['D18'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D18'].font = Font(name='times new roman', size=11)
            ws['D18'].border = Border(bottom=Side(border_style="thin"))
            if prof.discapacidad_neurologica:
                ws['D18'] = "Si"
            else:
                ws['D18'] = "No"
            ws['B19'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B19'].font = Font(name='times new roman', size=12, bold=True)
            ws['B19'] = 'Enfermedad cognitiva:'
            ws['D19'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D19'].font = Font(name='times new roman', size=11)
            ws['D19'].border = Border(bottom=Side(border_style="thin"))
            if prof.enfermedad_congenita:
                ws['D19'] = "Si"
            else:
                ws['D19'] = "No"
            ws['B20'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B20'].font = Font(name='times new roman', size=12, bold=True)
            ws['B20'] = 'Enfermedad respiratoria:'
            ws['D20'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D20'].font = Font(name='times new roman', size=11)
            ws['D20'].border = Border(bottom=Side(border_style="thin"))
            if prof.enfermedad_respiratoria:
                ws['D20'] = "Si"
            else:
                ws['D20'] = "No"
            ws['B21'].alignment = Alignment(horizontal="right", vertical="center")
            ws['B21'].font = Font(name='times new roman', size=12, bold=True)
            ws['B21'] = 'Atención psicologica:'
            ws['D21'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D21'].font = Font(name='times new roman', size=11)
            ws['D21'].border = Border(bottom=Side(border_style="thin"))
            if prof.atencion_psicologica:
                ws['D21'] = "Si"
            else:
                ws['D21'] = "No"
        # Sección para fusionar campos
        ws.merge_cells('B3:B8')
        ws.merge_cells('A1:I1')
        ws.merge_cells('C3:D3')
        ws.merge_cells('C4:D4')
        ws.merge_cells('C5:D5')
        ws.merge_cells('C6:D6')
        ws.merge_cells('C7:D7')
        ws.merge_cells('C8:D8')
        ws.merge_cells('E3:G3')
        ws.merge_cells('E4:G4')
        ws.merge_cells('E5:G5')
        ws.merge_cells('E6:G6')
        ws.merge_cells('E7:G7')
        ws.merge_cells('E8:G8')
        ws.merge_cells('B10:C10')
        ws.merge_cells('B11:C11')
        ws.merge_cells('B12:C12')
        ws.merge_cells('B13:C13')
        ws.merge_cells('D13:E13')
        ws.merge_cells('F10:H10')
        ws.merge_cells('B15:H15')
        ws.merge_cells('B17:C17')
        ws.merge_cells('B18:C18')
        ws.merge_cells('B19:C19')
        ws.merge_cells('B20:C20')
        ws.merge_cells('B21:C21')
        ws.merge_cells('E17:F17')
        ws.merge_cells('E18:F18')
        ws.merge_cells('E19:F19')
        ws.merge_cells('E20:F20')
        ws.merge_cells('E21:F21')
        ws.merge_cells('A22:I22')
        ws.merge_cells('A2:A21')
        ws.merge_cells('I2:I21')
        # Tamaño de las Filas
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 20
        ws.row_dimensions[7].height = 20
        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 10
        ws.row_dimensions[10].height = 20
        ws.row_dimensions[11].height = 20
        ws.row_dimensions[12].height = 20
        ws.row_dimensions[13].height = 20
        ws.row_dimensions[14].height = 10
        ws.row_dimensions[15].height = 20
        ws.row_dimensions[16].height = 10
        ws.row_dimensions[17].height = 20
        ws.row_dimensions[18].height = 20
        ws.row_dimensions[19].height = 20
        ws.row_dimensions[20].height = 20
        ws.row_dimensions[21].height = 20
        ws.row_dimensions[22].height = 10
        # Color de las hojas
        ws.sheet_properties.tabColor = "1072BA"
        # Tamaño de las columnas
        ws.column_dimensions['A'].width = 1
        ws.column_dimensions['B'].width = 15.5
        ws.column_dimensions['C'].width = 9
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 1

        # Sección Historial de clases
        ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws1['A1'].font = Font(name='times new roman', size=14, bold=True)
        ws1['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws1['A1'] = 'Historial Académico'
        # Sección para fusionar campos
        ws1.merge_cells('A1:I1')
        # Tamaño de las columnas
        ws1.column_dimensions['A'].width = 1
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 10
        ws1.column_dimensions['D'].width = 10
        ws1.column_dimensions['E'].width = 10
        ws1.column_dimensions['F'].width = 10
        ws1.column_dimensions['G'].width = 10
        ws1.column_dimensions['H'].width = 10
        ws1.column_dimensions['I'].width = 1
        # Configuracionde descarga
        # establecer el nombre de mi archivo
        nombre_archivo = "ReporteProfesor.xlsx"
        # Definir tipo de respuesta que va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        # Se añade la imagen que se encuentra en la esquina de la hoja
        # ws.add_image(img, 'A1')
        # wb.save('logo-login.xlsx')
        wb.save(response)
        return response
